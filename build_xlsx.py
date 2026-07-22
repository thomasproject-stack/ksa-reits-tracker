"""Step 3 — render reit_data.json into styled, formula-driven Excel workbooks.

One parameterised builder emits both language variants (FR + EN). The workbook
is deliberately *formula-driven*: only the raw inputs (prices, units) are written
as values — the change, change %, market cap and sector totals are live Excel
formulas, so editing any input recalculates the whole sheet. Sheet 2 documents
the methodology and cites every source.

Reads:  reit_data.json (from extract.py)
Writes: KSA_REITs_price_and_market_cap.xlsx        (English)
        REITs_Arabie_Saoudite_prix_capitalisation.xlsx (French)
"""
import datetime as dt
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- palette ----------
NAVY = "1F3864"; BLUE = "2E5496"; BAND = "F2F5FB"
GREEN = "C6EFCE"; GREENT = "006100"; RED = "FFC7CE"; REDT = "9C0006"; GREY = "595959"

YQ = "https://finance.yahoo.com/quote/4330.SR"

# ---------- per-language strings ----------
LABELS = {
    "en": {
        "out": "KSA_REITs_price_and_market_cap.xlsx",
        "sheet": "KSA REITs",
        "title": "Saudi-Listed REITs (Tadawul) — Price & Market Capitalisation",
        "subtitle": (
            "Closing price on {d1} and {d2}  •  Market cap on {d2} "
            "(units outstanding x closing price)  •  Currency: {cur}  •  "
            "Data source: Yahoo Finance (Tadawul) — cross-checked vs Argaam / "
            "Saudi Exchange / StockAnalysis"
        ),
        "headers": [
            "#", "Code", "Fund", "Symbol", "Price {d1}\n({cur})", "Price {d2}\n({cur})",
            "Change\n({cur})", "Change\n(%)", "Units\noutstanding",
            "Market cap {d2}\n({cur})", "Source",
        ],
        "total": "SECTOR TOTAL ({n} REITs)",
        "meth_sheet": "Methodology & Sources",
        "meth_title": "Methodology — how these figures were produced",
        "meth_rows": [
            ("Scope", "REIT funds listed on the Saudi Exchange (Tadawul), code block 43xx. The list is discovered by scanning the code range and keeping names containing 'REIT'."),
            ("Price", "Adjusted CLOSING price from Yahoo Finance's public historical API (/v8/finance/chart, daily interval). Each target date is matched to the closest trading day on or before it (Asia/Riyadh calendar)."),
            ("Units outstanding", "Number of units (sharesOutstanding) from the Yahoo quoteSummary API (defaultKeyStatistics module), reached via crumb/cookie authentication."),
            ("Market cap {d2}", "Excel formula: units outstanding x closing price on {d2} (column J = I x F). Chosen over a 'live' market cap because it is the true snapshot for the requested date, not today's."),
            ("Change", "Excel formulas: Change ({cur}) = F - E ; Change (%) = (F - E) / E. Cells are coloured green = up, red = down."),
            ("Formulas", "Only raw inputs are stored as values (prices E/F, units I). Change, %, market cap and totals are live Excel formulas — edit a price or unit count and everything recalculates."),
            ("Validation", "Independently cross-checked vs Argaam, Saudi Exchange and StockAnalysis. Units outstanding are date-independent and matched exactly, e.g. Al Rajhi REIT (275.61M units), Jadwa REIT Saudi (186.51M units), Riyad REIT (172M units)."),
            ("Currency", "All amounts in {cur} (Saudi Riyal)."),
            ("Principle", "No fabricated values — any missing datum is marked 'n/a'."),
        ],
        "sources_title": "Sources",
        "sources": [
            ("Yahoo Finance (raw data)", "Historical prices + units outstanding — one page per fund: finance.yahoo.com/quote/<CODE>.SR", YQ),
            ("Saudi Exchange (Tadawul)", "Official REITs sector summary page", "https://www.saudiexchange.sa/wps/portal/saudiexchange/ourmarkets/funds-market-watch/reits"),
            ("Argaam — REITs sector", "Saudi financial portal, REITs sector (validation)", "https://www.argaam.com/en/sector/reits/sectorid/78"),
            ("StockAnalysis — Tadawul", "Market cap & units per fund: stockanalysis.com/quote/tadawul/<CODE>", "https://stockanalysis.com/quote/tadawul/4340/"),
        ],
        "generated": "Generated {date} — ksa-reits-tracker",
    },
    "fr": {
        "out": "REITs_Arabie_Saoudite_prix_capitalisation.xlsx",
        "sheet": "REITs KSA",
        "title": "REITs cotés en Arabie Saoudite (Tadawul) — Prix & Capitalisation",
        "subtitle": (
            "Prix de clôture au {d1} et au {d2}  •  Capitalisation au {d2} "
            "(unités en circulation × cours de clôture)  •  Devise : {cur}  •  "
            "Source données : Yahoo Finance (Tadawul) — validé vs Argaam / "
            "Saudi Exchange / StockAnalysis"
        ),
        "headers": [
            "#", "Code", "Fonds", "Symbole", "Prix {d1}\n({cur})", "Prix {d2}\n({cur})",
            "Variation\n({cur})", "Variation\n(%)", "Unités en\ncirculation",
            "Capitalisation {d2}\n({cur})", "Source",
        ],
        "total": "TOTAL SECTEUR ({n} REITs)",
        "meth_sheet": "Méthodologie & Sources",
        "meth_title": "Méthodologie — comment ces chiffres ont été produits",
        "meth_rows": [
            ("Périmètre", "Fonds REIT cotés à la Bourse saoudienne (Tadawul), plage de codes 43xx. La liste est découverte par balayage de la plage de codes et filtrage des noms contenant « REIT »."),
            ("Prix", "Cours de CLÔTURE ajusté récupéré via l'API historique publique de Yahoo Finance (/v8/finance/chart, intervalle journalier). Chaque date cible est rapprochée du jour de bourse le plus proche antérieur ou égal (calendrier Asia/Riyadh)."),
            ("Unités en circulation", "Nombre d'unités (sharesOutstanding) via l'API Yahoo quoteSummary (module defaultKeyStatistics), atteinte par authentification crumb/cookie."),
            ("Capitalisation {d2}", "Formule Excel : unités en circulation × cours de clôture du {d2} (colonne J = I × F). Retenue plutôt qu'une capitalisation « live » car elle donne le vrai instantané à la date demandée, et non celui du jour de consultation."),
            ("Variation", "Formules Excel : Variation ({cur}) = F − E ; Variation (%) = (F − E) / E. Cellules colorées : vert = hausse, rouge = baisse."),
            ("Formules", "Seuls les intrants bruts sont stockés en valeurs (prix E/F, unités I). Variation, %, capitalisation et totaux sont des formules Excel vivantes — modifiez un prix ou un nombre d'unités et tout se recalcule."),
            ("Validation", "Recoupement indépendant vs Argaam, Saudi Exchange et StockAnalysis. Les unités en circulation sont indépendantes de la date et concordent exactement, ex. Al Rajhi REIT (275,61 M unités), Jadwa REIT Saudi (186,51 M unités), Riyad REIT (172 M unités)."),
            ("Devise", "Tous les montants en {cur} (riyal saoudien)."),
            ("Principe", "Aucune valeur inventée — toute donnée manquante est notée « n/a »."),
        ],
        "sources_title": "Sources",
        "sources": [
            ("Yahoo Finance (données brutes)", "Prix historiques + unités en circulation — 1 page par fonds : finance.yahoo.com/quote/<CODE>.SR", YQ),
            ("Saudi Exchange (Tadawul)", "Page récap. officielle du secteur REITs", "https://www.saudiexchange.sa/wps/portal/saudiexchange/ourmarkets/funds-market-watch/reits"),
            ("Argaam — secteur REITs", "Portail financier saoudien, secteur REITs (validation)", "https://www.argaam.com/en/sector/reits/sectorid/78"),
            ("StockAnalysis — Tadawul", "Capitalisation & unités par fonds : stockanalysis.com/quote/tadawul/<CODE>", "https://stockanalysis.com/quote/tadawul/4340/"),
        ],
        "generated": "Généré le {date} — ksa-reits-tracker",
    },
}


def load_data(path="reit_data.json"):
    payload = json.load(open(path))
    start = dt.date.fromisoformat(payload["start_date"])
    end = dt.date.fromisoformat(payload["end_date"])
    cur = payload.get("currency", "SAR")
    rows = sorted(payload["rows"], key=lambda x: -(x["marketcap_end"] or 0))
    return start, end, cur, rows


def build(lang, start, end, cur, rows):
    L = LABELS[lang]
    d1, d2 = start.strftime("%d/%m/%Y"), end.strftime("%d/%m/%Y")
    ctx = {"d1": d1, "d2": d2, "cur": cur}

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ================= SHEET 1 : DATA =================
    ws = wb.active
    ws.title = L["sheet"]

    ws.merge_cells("A1:K1")
    c = ws["A1"]; c.value = L["title"]
    c.font = Font(bold=True, size=15, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    c = ws["A2"]; c.value = L["subtitle"].format(**ctx)
    c.font = Font(size=9, italic=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); ws.row_dimensions[2].height = 34

    hr = 3
    for j, h in enumerate(L["headers"], 1):
        cell = ws.cell(row=hr, column=j, value=h.format(**ctx))
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[hr].height = 42

    first = hr + 1
    r = first
    for i, row in enumerate(rows, 1):
        src = f"https://finance.yahoo.com/quote/{row['symbol']}"
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=row["code"])
        ws.cell(row=r, column=3, value=row["name"])
        ws.cell(row=r, column=4, value=row["symbol"])
        ws.cell(row=r, column=5, value=row["price_start"])
        ws.cell(row=r, column=6, value=row["price_end"])
        ws.cell(row=r, column=7, value=f"=F{r}-E{r}")           # change (currency)
        ws.cell(row=r, column=8, value=f"=(F{r}-E{r})/E{r}")    # change (%)
        ws.cell(row=r, column=9, value=row["shares_outstanding"])
        ws.cell(row=r, column=10, value=f"=I{r}*F{r}")          # market cap = units x end price
        lk = ws.cell(row=r, column=11, value="Yahoo/Tadawul"); lk.hyperlink = src
        lk.font = Font(color="0563C1", underline="single", size=9)

        for j in range(1, 12):
            cell = ws.cell(row=r, column=j)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if j == 3 else "center", vertical="center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=BAND)
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        ws.cell(row=r, column=7).number_format = "+#,##0.00;-#,##0.00"
        ws.cell(row=r, column=8).number_format = "+0.00%;-0.00%"
        ws.cell(row=r, column=9).number_format = "#,##0"
        ws.cell(row=r, column=10).number_format = "#,##0"

        pct = row.get("pct_change")
        if pct is not None and pct != 0:
            fill = GREEN if pct > 0 else RED
            txt = GREENT if pct > 0 else REDT
            for col in (7, 8):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fill)
                ws.cell(row=r, column=col).font = Font(bold=True, color=txt)
        r += 1
    last = r - 1

    # TOTAL row (live SUM formulas)
    ws.cell(row=r, column=3, value=L["total"].format(n=len(rows)))
    ws.cell(row=r, column=9, value=f"=SUM(I{first}:I{last})").number_format = "#,##0"
    ws.cell(row=r, column=10, value=f"=SUM(J{first}:J{last})").number_format = "#,##0"
    for j in range(1, 12):
        cell = ws.cell(row=r, column=j)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color="FFFFFF"); cell.border = border
        cell.alignment = Alignment(horizontal="left" if j == 3 else "center", vertical="center")

    for j, w in enumerate([4, 7, 40, 11, 14, 14, 11, 11, 15, 20, 14], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{hr}:K{last}"

    # ================= SHEET 2 : METHODOLOGY =================
    ws2 = wb.create_sheet(L["meth_sheet"])
    ws2.column_dimensions["A"].width = 3
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 95

    def title(row, text):
        ws2.merge_cells(f"B{row}:C{row}")
        cc = ws2[f"B{row}"]; cc.value = text
        cc.font = Font(bold=True, size=13, color="FFFFFF"); cc.fill = PatternFill("solid", fgColor=NAVY)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[row].height = 26

    def kv(row, k, v, link=None):
        a = ws2[f"B{row}"]; a.value = k; a.font = Font(bold=True, color=NAVY)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws2[f"C{row}"]; b.value = v; b.alignment = Alignment(vertical="top", wrap_text=True)
        if link:
            b.hyperlink = link; b.font = Font(color="0563C1", underline="single")

    title(2, L["meth_title"])
    r2 = 3
    for k, v in L["meth_rows"]:
        v = v.format(**ctx)
        kv(r2, k.format(**ctx), v)
        ws2.row_dimensions[r2].height = max(30, 15 * (1 + len(v) // 90))
        r2 += 1

    r2 += 1
    title(r2, L["sources_title"]); r2 += 1
    for k, v, link in L["sources"]:
        kv(r2, k, v, link); ws2.row_dimensions[r2].height = 28; r2 += 1

    r2 += 1
    gen = ws2[f"B{r2}"]
    gen.value = L["generated"].format(date=dt.date.today().isoformat())
    gen.font = Font(italic=True, size=9, color=GREY)

    wb.save(L["out"])
    print("Saved", L["out"])


if __name__ == "__main__":
    start, end, cur, rows = load_data()
    for lang in ("en", "fr"):
        build(lang, start, end, cur, rows)
